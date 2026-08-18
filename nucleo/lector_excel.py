"""
============================================================
LECTOR DE LA MATRIZ EXCEL
============================================================

Lee la matriz PRIZMA (.xlsx), detecta programa y curso,
localiza dinamicamente la cabecera "Semana correspondiente"
y devuelve solo las actividades OVI y OVA.
"""

import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

import configuracion as cfg
from nucleo.registro import obtener_registro

log = obtener_registro("lector_excel")


# ============================================================
# NORMALIZACION DE TEXTO
# ============================================================

def normalizar_texto(valor) -> str:
    """Quita acentos, espacios sobrantes y pasa a mayusculas."""
    if valor is None:
        return ""
    texto = str(valor).strip()
    if not texto:
        return ""
    descompuesto = unicodedata.normalize("NFD", texto)
    sin_acentos = "".join(c for c in descompuesto if unicodedata.category(c) != "Mn")
    return " ".join(sin_acentos.upper().split())


def texto_plano(valor) -> str:
    """Devuelve el texto tal cual, sin espacios sobrantes."""
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


# ============================================================
# ACTIVIDAD
# ============================================================

@dataclass
class Actividad:
    """Una fila procesable de la matriz."""

    fila_excel: int
    semana: str
    unidad: str
    auxiliar: str
    nombre: str
    categoria: str
    tipo_recurso: str
    referencia: str
    estado: str

    programa: str = ""
    curso: str = ""

    # Se completa mas adelante
    extension_esperada: str = ""      # ".h5p" o ".pdf"
    archivo_asignado: str = ""
    resultado: str = ""
    observacion: str = ""

    @property
    def nombre_normalizado(self) -> str:
        return normalizar_texto(self.nombre)

    @property
    def es_h5p(self) -> bool:
        return self.extension_esperada == ".h5p"

    def a_diccionario(self) -> dict:
        return {
            "fila_excel": self.fila_excel,
            "semana": self.semana,
            "unidad": self.unidad,
            "nombre": self.nombre,
            "categoria": self.categoria,
            "tipo_recurso": self.tipo_recurso,
            "referencia": self.referencia,
            "extension_esperada": self.extension_esperada,
            "archivo_asignado": self.archivo_asignado,
            "resultado": self.resultado,
            "observacion": self.observacion,
        }


@dataclass
class LecturaMatriz:
    """Resultado completo de leer una matriz."""

    programa: str = ""
    curso: str = ""
    semestre: str = ""
    variante: str = ""
    fila_cabecera: int = 0
    actividades: list = field(default_factory=list)
    omitidas: int = 0
    categorias_omitidas: dict = field(default_factory=dict)

    @property
    def cantidad_ovi(self) -> int:
        return sum(1 for a in self.actividades if normalizar_texto(a.categoria) == "OVI")

    @property
    def cantidad_ova(self) -> int:
        return sum(1 for a in self.actividades if normalizar_texto(a.categoria) == "OVA")

    @property
    def h5p_esperados(self) -> int:
        return sum(1 for a in self.actividades if a.extension_esperada == ".h5p")

    @property
    def pdf_esperados(self) -> int:
        return sum(1 for a in self.actividades if a.extension_esperada == ".pdf")


# ============================================================
# DETECCION DE CABECERA
# ============================================================

def buscar_fila_cabecera(hoja) -> int:
    """
    Busca la fila cuya columna A contiene "Semana correspondiente".
    Devuelve 0 si no la encuentra.
    """
    limite = min(hoja.max_row, cfg.FILAS_MAXIMAS_BUSQUEDA_CABECERA)
    for numero_fila in range(1, limite + 1):
        valor = hoja.cell(row=numero_fila, column=cfg.COLUMNA_SEMANA).value
        if cfg.TEXTO_CABECERA_SEMANA in normalizar_texto(valor):
            log.info("Cabecera detectada en la fila %s", numero_fila)
            return numero_fila
    return 0


def leer_encabezado_curso(hoja) -> tuple:
    """
    Lee programa, curso, semestre y variante desde las primeras filas.
    Toma la primera celda con contenido de cada fila.
    """
    def primera_celda_con_texto(numero_fila: int) -> str:
        for columna in range(1, 9):
            valor = texto_plano(hoja.cell(row=numero_fila, column=columna).value)
            if valor:
                return valor
        return ""

    programa = primera_celda_con_texto(cfg.FILA_PROGRAMA)
    curso = primera_celda_con_texto(cfg.FILA_CURSO)
    semestre = primera_celda_con_texto(cfg.FILA_SEMESTRE)
    variante = primera_celda_con_texto(cfg.FILA_VARIANTE)

    return programa, curso, semestre, variante


# ============================================================
# CLASIFICACION DEL TIPO DE ARCHIVO
# ============================================================

def determinar_extension_esperada(categoria: str, tipo_recurso: str, referencia: str) -> str:
    """
    Reglas del proyecto:
      OVA -> siempre H5P
      OVI -> H5P o PDF segun tipo de recurso o referencia
      Infografia OVI -> H5P por ahora
    """
    categoria_norm = normalizar_texto(categoria)
    tipo_norm = normalizar_texto(tipo_recurso)
    referencia_norm = normalizar_texto(referencia)

    if categoria_norm == "OVA":
        return ".h5p"

    combinado = f"{tipo_norm} {referencia_norm}"

    if "H5P" in combinado:
        return ".h5p"
    if "INFOGRAFIA" in combinado:
        return ".h5p"
    if "PDF" in combinado:
        return ".pdf"

    # OVI sin senal clara: se asume H5P, que es el caso habitual
    return ".h5p"


def categoria_es_soportada(categoria: str) -> bool:
    categoria_norm = normalizar_texto(categoria)
    return categoria_norm in cfg.CATEGORIAS_SOPORTADAS


# ============================================================
# LECTURA PRINCIPAL
# ============================================================

def leer_actividades_excel(ruta_excel: Path) -> LecturaMatriz:
    """
    Lee la matriz completa y devuelve solo las actividades OVI y OVA.
    Las demas categorias se cuentan como omitidas.
    """
    ruta_excel = Path(ruta_excel)
    if not ruta_excel.exists():
        raise FileNotFoundError(f"No existe el archivo Excel: {ruta_excel}")

    log.info("Leyendo matriz: %s", ruta_excel.name)

    libro = load_workbook(ruta_excel, data_only=True, read_only=True)
    hoja = libro.active

    lectura = LecturaMatriz()
    lectura.programa, lectura.curso, lectura.semestre, lectura.variante = leer_encabezado_curso(hoja)

    fila_cabecera = buscar_fila_cabecera(hoja)
    if fila_cabecera == 0:
        libro.close()
        raise ValueError(
            f'No se encontro la cabecera "{cfg.TEXTO_CABECERA_SEMANA}" '
            f"en las primeras {cfg.FILAS_MAXIMAS_BUSQUEDA_CABECERA} filas."
        )
    lectura.fila_cabecera = fila_cabecera

    semana_arrastrada = ""
    unidad_arrastrada = ""

    for numero_fila in range(fila_cabecera + 1, hoja.max_row + 1):
        semana = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_SEMANA).value)
        unidad = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_UNIDAD).value)
        auxiliar = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_AUXILIAR).value)
        nombre = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_ACTIVIDAD).value)
        categoria = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_CATEGORIA).value)
        tipo = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_TIPO).value)
        referencia = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_REFERENCIA).value)
        estado = texto_plano(hoja.cell(row=numero_fila, column=cfg.COLUMNA_ESTADO).value)

        # Las celdas combinadas dejan huecos: se arrastra el ultimo valor
        if semana:
            semana_arrastrada = semana
        if unidad:
            unidad_arrastrada = unidad

        if not nombre and not categoria:
            continue

        if not categoria_es_soportada(categoria):
            if categoria:
                clave = normalizar_texto(categoria)
                lectura.categorias_omitidas[clave] = lectura.categorias_omitidas.get(clave, 0) + 1
                lectura.omitidas += 1
            continue

        if not nombre:
            log.warning("Fila %s con categoria %s pero sin nombre. Se omite.", numero_fila, categoria)
            continue

        actividad = Actividad(
            fila_excel=numero_fila,
            semana=semana or semana_arrastrada,
            unidad=unidad or unidad_arrastrada,
            auxiliar=auxiliar,
            nombre=nombre,
            categoria=normalizar_texto(categoria),
            tipo_recurso=tipo,
            referencia=referencia,
            estado=estado,
            programa=lectura.programa,
            curso=lectura.curso,
        )
        actividad.extension_esperada = determinar_extension_esperada(categoria, tipo, referencia)
        lectura.actividades.append(actividad)

    libro.close()

    log.info(
        "Matriz leida: %s actividades procesables (%s OVI, %s OVA), %s omitidas",
        len(lectura.actividades),
        lectura.cantidad_ovi,
        lectura.cantidad_ova,
        lectura.omitidas,
    )

    return lectura
